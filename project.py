from tkinter import *
import openpyxl
from openpyxl import Workbook
import pymysql
from tkinter import messagebox


file_name = "enquiry_data.txt"
window = Tk()
window.title("Besant Technologies")
window.geometry('600x650')
window.configure(bg="light blue")
excel_file_path = "enquiry_data.xlsx"


def initialize_database():
    connection = pymysql.connect(host="localhost",user="root",passwd="",database="employee")
    if connection.open:
        print("Database connection successful")
        return connection
    else:
        print("Failed to connect to the database")
        return None
    

def initialize_excel():
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Enquiry Data"
        headers = ["Date", "Name", "Mobile No", "Alternate No", "Email ID", "Address","Course Interested", "Batch Preferred", "How You Came To Know Us","Are You Experienced/Fresher", "Contact Person", "Counselor","Fees", "Comment", "Enquiry", "Registration"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header).font = Font(bold=True)
        workbook.save(excel_file_path)


def initialize_note():
    try:
        file = open(file_name, "x")
        file.write(f"{'Date':<30}{'Name':<30}{'Mobile_No':<30}{'Alternate_No':<30}{'Email_ID':<30}{'Address':<30}{'Course_Interested':<30}{'Batch_Preferred':<30}{'How_You_Came_To_Know_Us':<30}{'Are_You_Experienced/Fresher':<30}{'Contact_Person':<30}{'Counselor':<30}{'Fees':<30}{'Comment':<30}{'Enquiry':<30}{'Registration':<30}\n")
        file.write("=========================================================================================================================================================================================================================================================================================================================\n")
        file.close()
    except FileExistsError:
        pass
def update_data_base(connection,data):
    cursor = connection.cursor()
    query = """INSERT INTO enquiry_data 
               (date, name, mobile_no, alternate_no, email_id, address, course_interested, 
               batch_preferred, how_you_came_to_know_us, experience_status, contact_person, 
               counselor, fees, comment, enquiry, registration) 
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
    cursor.execute(query, data)
    connection.commit()
    print("Data added to the database successfully!")

def getdata():
    if not (Date.get() and Name.get() and Mobile.get() and Alternate.get() and Email.get() and
            Address.get() and Course.get() and Batch.get() and HowKnow.get() and Experience.get() and
            Contact.get() and Counselor.get() and Fees.get() and Comment.get()):
        messagebox.showwarning("Input Error", "please fill in all the details.")
        return
    if varEnqui.get() == 1:
        enquiry_state = "Yes"
    else:
        enquiry_state = "No"

    if varReg.get() == 1:
        registration_state = "Yes"
    else:
        registration_state = "No"

    file = open(file_name, 'a')
    file.write(f"{Date.get():<30}{Name.get():<30}{Mobile.get():<30}{Alternate.get():<30}{Email.get():<30}{Address.get():<30}{Course.get():<30}{Batch.get():<30}{HowKnow.get():<30}{Experience.get():<30}{Contact.get():<30}{Counselor.get():<30}{Fees.get():<30}{Comment.get():<30}{enquiry_state:<30}{registration_state:<30}\n")
    file.close()

    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    sheet.append([Date.get(), Name.get(), Mobile.get(), Alternate.get(), Email.get(),Address.get(), Course.get(), Batch.get(), HowKnow.get(), Experience.get(),Contact.get(), Counselor.get(), Fees.get(), Comment.get(),enquiry_state, registration_state])
    workbook.save(excel_file_path)

    if database_connection:
        update_data_base(database_connection, [Date.get(), Name.get(), Mobile.get(), Alternate.get(), Email.get(),Address.get(), Course.get(), Batch.get(), HowKnow.get(),Experience.get(), Contact.get(), Counselor.get(), Fees.get(),Comment.get(), enquiry_state, registration_state])
    else:
        print("No database connection. Data not saved to the database.")

    messagebox.showinfo("Submission Successful", "Data submitted successfully!")
    print("Data Submitted!")

    print("Data Submitted!")
    print("Date:", Date.get())
    print("Name:", Name.get())
    print("Mobile No:", Mobile.get())
    print("Alternate No:", Alternate.get())
    print("Email ID:", Email.get())
    print("Address:", Address.get())
    print("Course Interested:", Course.get())
    print("Batch Preferred:", Batch.get())
    print("How You Came To Know Us:", HowKnow.get())
    print("Are You Experienced or Fresher:", Experience.get())
    print("Contact Person-Besant Technology:", Contact.get())
    print("Counselor:", Counselor.get())
    print("Fees:", Fees.get())
    print("Comment:", Comment.get())
    print("Enquiry:", enquiry_state)
    print("Registration:", registration_state)



Label(window, text=" Technology Enquiry Form", font="Ariel 20 bold", fg="red", bg="light blue").grid(row=0, column=0, columnspan=2, pady=20)
Label(window, text="Date:", font=15, bg="light blue").grid(row=1, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Name:", font=15, bg="light blue").grid(row=2, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Mobile No:", font=15, bg="light blue").grid(row=3, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Alternate No:", font=15, bg="light blue").grid(row=4, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Email ID:", font=15, bg="light blue").grid(row=5, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Address:", font=15, bg="light blue").grid(row=6, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Course Interested:", font=15, bg="light blue").grid(row=7, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Batch Preferred:", font=15, bg="light blue").grid(row=8, column=0, padx=10, pady=5, sticky="w")
Label(window, text="How You Came To Know Us:", font=15, bg="light blue").grid(row=9, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Are You Experienced or Fresher:", font=15, bg="light blue").grid(row=10, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Contact Person-Besant Technology:", font=15, bg="light blue").grid(row=11, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Counselor:", font=15, bg="light blue").grid(row=12, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Fees:", font=15, bg="light blue").grid(row=13, column=0, padx=10, pady=5, sticky="w")
Label(window, text="Comment:", font=15, bg="light blue").grid(row=14, column=0, padx=10, pady=5, sticky="w")

Date = Entry(window, width=30, bd=2, font=15)
Date.grid(row=1, column=1, padx=10, pady=5)
Name = Entry(window, width=30, bd=2, font=15)
Name.grid(row=2, column=1, padx=10, pady=5)
Mobile = Entry(window, width=30, bd=2, font=15)
Mobile.grid(row=3, column=1, padx=10, pady=5)
Alternate = Entry(window, width=30, bd=2, font=15)
Alternate.grid(row=4, column=1, padx=10, pady=5)
Email = Entry(window, width=30, bd=2, font=15)
Email.grid(row=5, column=1, padx=10, pady=5)
Address = Entry(window, width=30, bd=2, font=15)
Address.grid(row=6, column=1, padx=10, pady=5)
Course = Entry(window, width=30, bd=2, font=15)
Course.grid(row=7, column=1, padx=10, pady=5)
Batch = Entry(window, width=30, bd=2, font=15)
Batch.grid(row=8, column=1, padx=10, pady=5)
HowKnow = Entry(window, width=30, bd=2, font=15)
HowKnow.grid(row=9, column=1, padx=10, pady=5)
Experience = Entry(window, width=30, bd=2, font=15)
Experience.grid(row=10, column=1, padx=10, pady=5)
Contact = Entry(window, width=30, bd=2, font=15)
Contact.grid(row=11, column=1, padx=10, pady=5)
Counselor = Entry(window, width=30, bd=2, font=15)
Counselor.grid(row=12, column=1, padx=10, pady=5)
Fees = Entry(window, width=30, bd=2, font=15)
Fees.grid(row=13, column=1, padx=10, pady=5)
Comment = Entry(window, width=30, bd=2, font=15)
Comment.grid(row=14, column=1, padx=10, pady=5)

def confirm_exit():
    if messagebox.askyesno("Exit Confirmation", "are you sure you want to exit"):
        window.destroy()

varEnqui = IntVar()
varReg = IntVar()
Checkbutton(window, text="Enquiry", variable=varEnqui, bg="light blue", font=15).grid(row=15, column=0, sticky="e")
Checkbutton(window, text="Registration", variable=varReg, bg="light blue", font=15).grid(row=15, column=1, sticky="w")

Button(window, text="Submit", bg="green", fg="white", font=15, width=10, command=getdata).grid(row=16, column=0, padx=10, pady=20, sticky="e")
Button(window, text="Quit", bg="red", fg="white", font=15, width=10, command=confirm_exit).grid(row=16, column=1, padx=10, pady=20, sticky="w")


database_connection = initialize_database()
initialize_excel()
initialize_note()
window.mainloop()
